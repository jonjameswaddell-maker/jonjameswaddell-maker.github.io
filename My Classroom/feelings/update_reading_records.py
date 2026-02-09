import openpyxl
from openpyxl.styles import Alignment
from copy import copy
import re
import string

def clean_and_format_text(text):
    """
    Parses the cell content to separate Book Name and Level.
    Capitalizes the Book Name and ensures a newline separator.
    Returns: (formatted_text, extracted_level_string)
    """
    if not text:
        return None, None
    
    val = str(text).strip()
    
    # Regex strategy:
    # Look for a level number (digits, optional '+') that appears 
    # either after a newline, at the end of the string, or separated by space.
    # We prioritize finding the last numeric token which usually indicates the level.
    
    # Pattern explanation:
    # [\s\n]  : matches whitespace or newline before the level
    # (\d{1,2}\+?) : matches 1-2 digits and an optional plus sign (Group 1)
    # (?=\s|$|\n) : lookahead to ensure it ends with space, newline or end of string
    matches = list(re.finditer(r'[\s\n]+(\d{1,2}\+?)(?=\s|$|\n)', val))
    
    # Fallback: if no space separator found, look for number at very end of string
    if not matches:
        matches = list(re.finditer(r'(\d{1,2}\+?)$', val))

    if matches:
        # Use the last match found as the level
        m = matches[-1]
        level_str = m.group(1)
        
        # Split the string into parts
        # Part 1: Book Name (everything before the level match)
        name_part = val[:m.start()].strip()
        
        # Part 2: Suffix (everything after the level match, e.g. "(read with me)")
        suffix_part = val[m.end():].strip()
        
        # Capitalize the Book Name (Title Case)
        # string.capwords is good for "a day in london" -> "A Day In London"
        name_part = string.capwords(name_part)
        
        # Reconstruct the string: Name \n Level
        new_val = f"{name_part}\n{level_str}"
        
        # If there was a note/suffix, append it on a new line
        if suffix_part:
            new_val += f"\n{suffix_part}"
            
        return new_val, level_str
        
    return val, None

def update_excel_sheet(filename):
    print(f"Loading {filename}...")
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print("Error: File not found. Please ensure the Excel file is in the same folder.")
        return

    sheet_name = "2B Liben 2.0"
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found.")
        return
        
    ws = wb[sheet_name]
    
    # --- Step 1: Learn the Color Codes from C57:C77 ---
    print("Reading color legend from C57:C77...")
    level_color_map = {}
    
    # Iterate through the legend rows
    for row in range(57, 78):
        # Column C is index 3
        cell = ws.cell(row=row, column=3)
        level_val = str(cell.value).strip() if cell.value is not None else None
        
        if level_val:
            # Store the fill style object
            # We map the string representation of the level (e.g., "8", "1+") to the fill
            level_color_map[level_val] = cell.fill

    print(f"Found color codes for levels: {list(level_color_map.keys())}")

    # --- Step 2: Process the Main Table ---
    # We iterate from row 2 down to 56 (stopping before the legend)
    # We iterate from column 6 (F) to the max column
    
    print("Processing table rows 2-56...")
    
    # Define columns to skip (A=1, B=2, C=3, D=4, E=5)
    start_col_idx = 6 
    
    for row_idx in range(2, 57):
        for col_idx in range(start_col_idx, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.value:
                original_text = cell.value
                new_text, extracted_level = clean_and_format_text(original_text)
                
                # Update text formatting
                cell.value = new_text
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                # Apply color if level is found in our map
                if extracted_level and extracted_level in level_color_map:
                    # We copy the fill pattern from the legend to the current cell
                    cell.fill = copy(level_color_map[extracted_level])

    # --- Step 3: Save ---
    output_filename = "Liben_Kbely_Reading_records_Updated.xlsx"
    wb.save(output_filename)
    print(f"Done! Saved updated file as: {output_filename}")

if __name__ == "__main__":
    # Assuming the file name matches the one uploaded
    file_path = "Liben Kbely Reading records (2).xlsx" 
    update_excel_sheet(file_path)